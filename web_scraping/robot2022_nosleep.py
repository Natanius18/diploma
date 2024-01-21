from datetime import datetime
from time import sleep

from openpyxl import load_workbook
from selenium.webdriver import Chrome, Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select


def check_and_close_error_message():
    if browser.find_elements(By.CLASS_NAME, 'ui-dialog-buttonset'):
        # sleep(2)
        browser.find_element(By.CLASS_NAME, 'ui-button').click()


def close_last_tab():
    browser.close()
    browser.switch_to.window(window_name=browser.window_handles[0])


def find_derj_value(offer):
    if offer.find_elements(By.CLASS_NAME, 'offer-max-order'):
        return offer.find_element(By.CLASS_NAME, 'offer-max-order').find_element(By.TAG_NAME, 'dd').text
    elif offer.find_elements(By.CLASS_NAME, 'offer-order-budget'):
        return offer.find_element(By.CLASS_NAME, 'offer-order-budget').find_element(By.TAG_NAME, 'dd').text
    else:
        return '-'


def read_info(offer):
    speciality = offer.find_element(By.CLASS_NAME, 'offer-university-specialities-name').find_element(By.TAG_NAME,
                                                                                                      'dd').text
    specialization = ""
    for i in range(0, len(speciality)):
        if speciality[i].islower():
            specialization = specialization.join(speciality[i:])
            speciality = speciality[:i - 1]
            break
    max_order = find_derj_value(offer)

    if offer.find_elements(By.CLASS_NAME, 'stats-field-km'):
        min = offer.find_element(By.CLASS_NAME, 'stats-field-km').find_element(By.CLASS_NAME, 'value').text
    elif offer.find_elements(By.CLASS_NAME, 'stats-field-rm'):
        min = offer.find_element(By.CLASS_NAME, 'stats-field-rm').find_element(By.CLASS_NAME, 'value').text
    else:
        min = '-'

    if offer.find_elements(By.CLASS_NAME, 'stats-field-kx'):
        max = offer.find_element(By.CLASS_NAME, 'stats-field-kx').find_element(By.CLASS_NAME, 'value').text
    else:
        max = '-'

    if offer.find_elements(By.CLASS_NAME, 'stats-field-ka'):
        avg = offer.find_element(By.CLASS_NAME, 'stats-field-ka').find_element(By.CLASS_NAME, 'value').text
    else:
        avg = '-'

    education_form = offer.find_element(By.CLASS_NAME, 'offer-education-form-name').find_element(By.TAG_NAME, 'dd').text

    if offer.find_elements(By.CLASS_NAME, 'offer-regional-coeff'):
        regional_coeff = offer.find_element(By.CLASS_NAME, 'offer-regional-coeff').find_element(By.TAG_NAME, 'dd').text
    else:
        regional_coeff = "-"

    qualification = offer.parent.find_element(By.CLASS_NAME, 'university-offers-title').find_element(By.TAG_NAME,
                                                                                                     'h6').text

    if offer.find_elements(By.CLASS_NAME, 'offer-order-license'):
        license_value = offer.find_element(By.CLASS_NAME, 'offer-order-license').find_element(By.TAG_NAME, 'dd').text
    else:
        license_value = "-"

    # print("uni_name:", uni_name,
    #       "\nspeciality:", speciality,
    #       "\nqualification:", qualification,
    #       "\nspecialization:", specialization,
    #       "\neducation_form:", education_form,
    #       "\nmax_order:", max_order,
    #       "\navg_konkusrs_value:", avg_konkurs_value,
    #       "\nlicense_value:", license_value,
    #       "\nregional_coeff:", regional_coeff, "\n\n\n")

    # sleep(1)
    # close_last_tab()

    ws.append([uni_name, speciality, specialization, education_form, max_order, avg, min, max, license_value,
               qualification, regional_coeff])
    ws["Q1"] = button_num
    print('spec_num_start:', spec_num + 1, '/', len(specialities), '\t\tcity_num_start:',
          # city_num + 1, '/', len(cities),
          '\t\tbutton:', button_num + 1, '/', len(buttons))
    wb.save(filename)
    # sleep(6)


def scrap_page():
    global button_num, uni_name, spec_num, button_num_start, x

    if spec_num > spec_num_start:  # or city_num > city_num_start:
        button_num_start = 0

    for button_num in range(button_num_start, len(buttons)):
        ws["Q1"] = button_num
        wb.save(filename)

        button = buttons[button_num]
        browser.implicitly_wait(1)

        actions.move_to_element(button).perform()
        # sleep(3)
        button.click()
        # print("button:", button_num, "/", len(buttons))
        # sleep(2)
        offer_path = "//div[@class='university'][" + (button_num + 1).__str__() + "]"
        offers = browser.find_element(By.XPATH, offer_path).find_elements(By.CLASS_NAME, 'offer')
        if not offers:
            print('тут пусто')
            # sleep(5)
            continue
        uni_name = browser.find_element(By.XPATH, offer_path).find_element(By.TAG_NAME, 'h5').text
        browser.implicitly_wait(0)
        # sleep(3)
        for offer_num in range(len(offers)):
            offer = offers[offer_num]
            x = x + 1

            if offer.find_elements(By.CLASS_NAME, 'offer-offer-type-name'):
                offer_type = offer.find_element(By.CLASS_NAME, 'offer-offer-type-name').find_element(By.TAG_NAME, 'dd').text
                if offer_type.__contains__('з пріоритетом'):
                    # actions.move_to_element(offer).perform()
                    # sleep(7)
                    # offer.find_element(By.CLASS_NAME, 'button').click()
                    # sleep(4)
                    read_info(offer)
                    print('Просмотрел уже', x, 'офферов', datetime.now())

        # check_and_close_error_message()
        # actions.move_to_element(button).perform()
        # sleep(1)
        # button.click()
        # sleep(1)


filename = "data2022.xlsx"
x = 0
while True:
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_experimental_option("detach", True)

        browser = Chrome(options=chrome_options)
        browser.maximize_window()
        actions = ActionChains(browser)
        browser.implicitly_wait(1)

        wb = load_workbook(filename)
        ws = wb.active
        ws.append(["Університет", "Спеціальність", "Спеціалізація", "Форма навчання", "Макс. обсяг держзамовлення",
                   "СЕР", "МІН", "МАКС", "Ліцензійний обсяг",
                   "Точно бакалавр?", "Регіональний коефіцієнт"])

        url = 'https://vstup2022.edbo.gov.ua/offers/'
        license_path = "//*[contains(text(), 'Ліцензійний обсяг:')]"

        browser.get(url)

        choose_speciality = Select(browser.find_element(By.ID, 'offers-search-speciality'))
        specialities = []
        for spec in browser.find_elements(By.CLASS_NAME, 'level-1'):
            specialities.append(spec.text)

        choose_city = Select(browser.find_element(By.ID, 'offers-search-region'))
        # cities = []
        # for city in browser.find_element(By.ID, 'offers-search-region').find_elements(By.TAG_NAME, 'option'):
        #     cities.append(city.text)

        spec_num_start = ws["O1"].value  # менять тут!! -----------------------------------------------------------
        city_num_start = ws["P1"].value  # менять тут!! ------------------------------------------------------------
        button_num_start = ws["Q1"].value

        for spec_num in range(spec_num_start, len(specialities)):
            choose_speciality.select_by_visible_text(specialities[spec_num])
            # choose_speciality.select_by_value('014')
            if spec_num > spec_num_start:
                city_num_start = 0
            # for city_num in range(city_num_start, len(cities)):
            #     choose_city.select_by_visible_text(cities[city_num])
            #     # sleep(1)
            but_search = browser.find_element(By.ID, 'offers-search-button')
            actions.move_to_element(but_search)
            but_search.click()

            browser.implicitly_wait(3)
            if browser.find_elements(By.CLASS_NAME, 'close'):
                print("по специальности", specialities[spec_num], "ничего не нашел")
                ws["O1"] = spec_num + 1
                wb.save(filename)
                browser.find_element(By.CLASS_NAME, 'close').click()
                sleep(1)
                continue
            ws["O1"] = spec_num
            wb.save(filename)
            buttons = browser.find_elements(By.CLASS_NAME, 'university-title')
            scrap_page()
        print("2021 FINISH!")
        break
    except Exception as e:
        wb.save(filename)
        err_path = "//*[contains(text(), 'перевищили')]"
        if browser.find_elements(By.XPATH, err_path):
            print("Опять превышаем")
        else:
            print('Error occurred:', e)
        sleep(3)
        browser.quit()
