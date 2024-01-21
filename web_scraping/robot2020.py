import time
from datetime import datetime
from time import sleep

import winsound
from selenium.webdriver import Chrome, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select


def check_and_close_error_message():
    if browser.find_elements(By.CLASS_NAME, 'ui-dialog-buttonset'):
        sleep(2)
        browser.find_element(By.CLASS_NAME, 'ui-button').click()


def close_last_tab():
    browser.close()
    browser.switch_to.window(window_name=browser.window_handles[0])


def find_derj_value():
    if browser.find_elements(By.CLASS_NAME, 'max-order'):
        return browser.find_element(By.CLASS_NAME, 'max-order').find_element(By.TAG_NAME, 'div').text
    else:
        elements = browser.find_element(By.ID, 'offer-wrapper').find_elements(By.TAG_NAME, 'label')
        for e in elements:
            if e.text.__contains__("держзамовлення"):
                return e.find_element(By.XPATH, "..").find_element(By.TAG_NAME, 'div').text


def read_info():
    browser.switch_to.window(window_name=browser.window_handles[-1])
    speciality = browser.find_element(By.CLASS_NAME, 'speciality-name').text
    max_order = find_derj_value()
    if browser.find_elements(By.CLASS_NAME, 'avg-konkurs-value'):
        avg_konkurs_value = browser.find_element(By.CLASS_NAME, 'avg-konkurs-value').find_element(By.TAG_NAME, 'div').text
    else:
        avg_konkurs_value = "-"
    if browser.find_elements(By.CLASS_NAME, 'specialization'):
        specialization = browser.find_element(By.CLASS_NAME, 'specialization').find_element(By.TAG_NAME, 'div').text
    else:
        specialization = "-"
    education_form = browser.find_element(By.CLASS_NAME, 'education-form').find_element(By.TAG_NAME, 'div').text
    if browser.find_elements(By.CLASS_NAME, 'regional-coeff'):
        regional_coeff = browser.find_element(By.CLASS_NAME, 'regional-coeff').find_element(By.TAG_NAME, 'div').text
    else:
        regional_coeff = "-"
    qualification = browser.find_element(By.CLASS_NAME, 'qualification').text
    if browser.find_elements(By.XPATH, license_path):
        license_value = browser.find_element(By.XPATH, license_path).find_element(By.XPATH, '..').find_element(By.TAG_NAME, 'div').text
    else:
        license_value = "-"

    print("uni_name:", uni_name,
          "\nspeciality:", speciality,
          "\nqualification:", qualification,
          "\nspecialization:", specialization,
          "\neducation_form:", education_form,
          "\nmax_order:", max_order,
          "\navg_konkusrs_value:", avg_konkurs_value,
          "\nlicense_value:", license_value,
          "\nregional_coeff:", regional_coeff, "\n\n\n")

    sleep(6)
    close_last_tab()

    ws.append([uni_name, speciality, specialization, education_form, max_order, avg_konkurs_value, license_value,
               qualification, regional_coeff, button_num])
    ws["Q1"] = button_num
    print('spec_num_start:', spec_num + 1, '/', len(specialities), '\t\tcity_num_start:',
          city_num + 1, '/', len(cities), '\t\tbutton:', button_num + 1, '/', len(buttons))
    wb.save(filename)
    sleep(6)


def scrap_page():
    global button_num, uni_name, spec_num, button_num_start, x

    if spec_num > spec_num_start or city_num > city_num_start:
        button_num_start = 0

    for button_num in range(button_num_start, len(buttons)):
        ws["Q1"] = button_num
        wb.save(filename)

        button = buttons[button_num]

        actions.move_to_element(button).perform()
        sleep(3)
        button.click()
        # print("button:", button_num, "/", len(buttons))
        sleep(2)
        offer_path = "//div[@class='offer-university'][" + (button_num + 1).__str__() + "]"
        offers = browser.find_element(By.XPATH, offer_path).find_elements(By.CLASS_NAME, 'offer-wrapper')
        if not offers:
            print('тут пусто')
            sleep(5)
            continue
        uni_name = browser.find_element(By.XPATH, offer_path).find_element(By.TAG_NAME, 'h2').text
        browser.implicitly_wait(0)
        sleep(3)
        for offer_num in range(len(offers)):
            offer = offers[offer_num]
            x = x + 1
            print('Просмотрел уже', x, 'офферов', datetime.now())

            if offer.find_elements(By.CLASS_NAME, 'offer-type'):
                offer_type = offer.find_element(By.CLASS_NAME, 'offer-type').find_element(By.TAG_NAME, 'div').text
                if offer_type.__contains__('із вказанням пріоритетності'):
                    actions.move_to_element(offer).perform()
                    sleep(7)
                    offer.find_element(By.CLASS_NAME, 'button').click()
                    sleep(4)
                    read_info()

        check_and_close_error_message()
        actions.move_to_element(button).perform()
        sleep(1)
        button.click()
        sleep(1)


filename = "data2020.xlsx"
x = 0
while True:
    try:
        chrome_options = Options()
        # chrome_options.add_argument("--headless")
        # chrome_options.add_experimental_option("detach", True)

        browser = Chrome(options=chrome_options)
        browser.maximize_window()
        actions = ActionChains(browser)
        browser.implicitly_wait(10)

        wb = load_workbook(filename)
        ws = wb.active
        ws.append(["Університет", "Спеціальність", "Спеціалізація", "Форма навчання", "Макс. обсяг держзамовлення", "Середні бали ЗНО", "Ліцензійний обсяг",
                   "Точно бакалавр?", "Регіональний коефіцієнт"])

        url = 'https://vstup2020.edbo.gov.ua/offers/'
        license_path = "//*[contains(text(), 'Ліцензійний обсяг:')]"

        browser.get(url)

        choose_speciality = browser.find_element(By.ID, 'select2-offers-search-speciality-container')
        choose_speciality.click()
        specialities = browser.find_elements(By.CLASS_NAME, 'select2-results__options')[0].text.split('\n')
        choose_speciality.click()

        choose_city = browser.find_element(By.ID, 'select2-offers-search-region-container')
        choose_city.click()
        cities = browser.find_elements(By.CLASS_NAME, 'select2-results__options')[0].text.split('\n')
        choose_city.click()
        # spec_num_start = ws["O1"].value      # менять тут!! -----------------------------------------------------------
        spec_num_start = 0      # менять тут!! -----------------------------------------------------------
        city_num_start = ws["P1"].value  # менять тут!! ------------------------------------------------------------
        button_num_start = ws["Q1"].value

        # for spec_num in range(spec_num_start, len(specialities)):
        for spec_num in range(1, 2):
            choose_speciality = browser.find_element(By.ID, 'select2-offers-search-speciality-container')
            choose_speciality.click()
            input_speciality = browser.find_element(By.CLASS_NAME, 'select2-search__field')
            input_speciality.send_keys('275')
            # input_speciality.send_keys(specialities[spec_num])
            input_speciality.send_keys(Keys.ENTER)
            if spec_num > spec_num_start:
                city_num_start = 0
            for city_num in range(city_num_start, len(cities)):
                choose_city.click()
                sleep(0.1)
                ddelement = Select(browser.find_element(By.ID, 'offers-search-region'))
                ddelement.select_by_visible_text(cities[city_num])
                choose_city.click()
                sleep(1)
                but_search = browser.find_element(By.ID, 'button-offers-search')
                actions.move_to_element(but_search)
                but_search.click()

                browser.implicitly_wait(5)
                if browser.find_elements(By.CLASS_NAME, 'ui-dialog-buttonset'):
                    print("по специальности", specialities[spec_num], "в городе", cities[city_num], "ничего не нашел")
                    sleep(0.5)
                    browser.find_element(By.CLASS_NAME, 'ui-button').click()
                    sleep(8)
                    continue
                ws["O1"] = spec_num
                wb.save(filename)
                buttons = browser.find_elements(By.CLASS_NAME, 'button-offer-university')
                scrap_page()
        print("2020 FINISH!")
    except Exception as e:
        wb.save(filename)
        err_path = "//*[contains(text(), 'перевищили')]"
        if browser.find_elements(By.XPATH, err_path):
            print("Опять превышаем")
        else:
            print('Error occurred:', e)
            for spec_num in range(2):
                # winsound.MessageBeep(winsound.SND_NOWAIT)
                sleep(1.5)
        sleep(60)
        browser.quit()
